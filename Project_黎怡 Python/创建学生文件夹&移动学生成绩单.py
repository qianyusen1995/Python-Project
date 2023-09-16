import os
import openpyxl
import shutil

# 打开源 Excel 文件
source_excel = openpyxl.load_workbook(r'C:\Users\QianYus\OneDrive\#Python0726\Github\Python-Project\Project_黎怡 Python\黎怡的班级.xlsx')
source_sheet = source_excel.active

# 获取源 Excel 的总行数
total_rows = source_sheet.max_row

# 遍历每一行，从第二行开始
for row_num in range(2, total_rows + 1): #左闭右开区间 所以要+1
    # 读取第三列的值，作为文件夹名字
    folder_name = source_sheet.cell(row=row_num, column=1).value

    # 检查文件夹是否已存在，如果不存在则创建
    if not os.path.exists(folder_name):
        os.mkdir(folder_name)

    # 构建要移动的文件的路径
    source_file = source_sheet.cell(row=row_num, column=1).value + '.xlsx'
    source_file_path = os.path.join(os.getcwd(), source_file)

    # 构建要移动到的目标文件夹的路径
    destination_folder = os.path.join(os.getcwd(), folder_name)

    # 移动文件到目标文件夹
    shutil.move(source_file_path, os.path.join(destination_folder, source_file))

# 关闭源 Excel 文件
source_excel.close()
