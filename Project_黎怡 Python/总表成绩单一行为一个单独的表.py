import openpyxl

# 打开源 Excel 文件
source_excel = openpyxl.load_workbook(r'C:\Users\QianYus\OneDrive\#Python0726\Github\Python-Project\Linkedin\七年级语文细分pro.xlsx')
source_sheet = source_excel.active

# 获取源 Excel 的总行数
total_rows = source_sheet.max_row

# 创建新的 Excel 文件
new_excel = openpyxl.Workbook()
new_sheet = new_excel.active

# 复制第一行到新 Excel
for row_num in range(1, 2):
    for col_num, cell in enumerate(source_sheet[row_num]):
        new_sheet.cell(row=row_num, column=col_num + 1, value=cell.value)

# 遍历每一行，从第二行开始创建新的 Excel 文件
for row_num in range(2, total_rows + 1):
    # 读取第三列的值，作为新 Excel 文件名
    new_file_name = source_sheet.cell(row=row_num, column=3).value + '.xlsx'

    # 复制当前行到新 Excel 的下一行
    for col_num, cell in enumerate(source_sheet[row_num]):
        new_sheet.cell(row=2, column=col_num + 1, value=cell.value)

    # 保存新 Excel 文件
    new_excel.save(new_file_name)

# 关闭源 Excel 文件
source_excel.close()
